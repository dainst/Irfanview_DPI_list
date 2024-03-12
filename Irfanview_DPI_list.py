#!/usr/bin/env python

__license__ = 'GPL'
__version__ = '0.9'

# ********** Setup
import subprocess
import os
import sys
import datetime
import requests

from distutils.spawn import find_executable

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

import tkinter as tk
from tkinter import filedialog, ttk
import tkinter.messagebox as messagebox

def check_version(__version__):
    url = 'https://fabfab1.github.io/Irfanview_DPI_list/i_dpi_list_ver.html'
    try:
        resp = requests.get(url)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Error", "No internet connection, push OK to proceed")
        return

    try:
        latest_version = resp.text.strip()  # Remove any leading/trailing whitespace
    except AttributeError:
        print('Warning: Unexpected response from server')
        return

    if __version__ != latest_version:
        if not messagebox.askokcancel("Error", "Current version is outdated. Push OK to proceed or EXIT to exit."):
            sys.exit(1)

def get_irfanview_path():
    global irfan_prog_cmd
    # Find and Run Irfanview
    def check_irfan_exists(irfan_prog_cmd):
        if not os.path.isfile(irfan_prog_cmd):
            irfan_prog_cmd = None
        return irfan_prog_cmd

    irfan_prog_name = 'i_view64.exe'
    irfan_prog_cmd = find_executable(irfan_prog_name)
    if not irfan_prog_cmd:
        irfan_prog_cmd = r'C:\Program Files\IrfanView\i_view64.exe'
        irfan_prog_cmd = check_irfan_exists(irfan_prog_cmd)
    if not irfan_prog_cmd:
        irfan_prog_cmd = os.path.join(os.path.expanduser('~'), r'C:\PortableApps\IrfanViewPortable\App\IrfanView64\i_view64.exe')
        irfan_prog_cmd = check_irfan_exists(irfan_prog_cmd)
    if not irfan_prog_cmd:
        irfan_prog_name = 'i_view32.exe'
        irfan_prog_cmd = find_executable(irfan_prog_name)
    if not irfan_prog_cmd:
        irfan_prog_cmd = os.path.join(os.path.expanduser('~'), r'C:\PortableApps\IrfanViewPortable\App\IrfanView\i_view32.exe')
        irfan_prog_cmd = check_irfan_exists(irfan_prog_cmd)
    if not irfan_prog_cmd:
        irfan_prog_cmd = r'C:\Program Files (x86)\IrfanView\i_view32.exe'
        irfan_prog_cmd = check_irfan_exists(irfan_prog_cmd)
    if not irfan_prog_cmd:
        raise Exception('Irfanview not installed, please install and run again./nIf installed, please check that the Irfanview_DPI_list program is running from your laptop harddrive/nand not a cloud drive.')

def generate_excel():
    global pic_dir
    global irfan_prog_cmd

    # Filenames
    irfan_info_txt = 'DPI_list_irfanviewOUT.txt'

    # This calls IrfanView and creates TXT file
    irfan_info_txt = os.path.join(pic_dir, irfan_info_txt)
    if os.path.exists(irfan_info_txt):  # Delete TXT file if it already exists
        os.remove(irfan_info_txt)
    irfan_prog_run = irfan_prog_cmd + ' ' + '"' + os.path.join(pic_dir, '*.*') + '"' + ' /silent /info=' + '"' + irfan_info_txt + '"'

    # Generate Excel Filename
    last_pic_dir_name = os.path.basename(os.path.normpath(pic_dir))
    excel_filename = '^' + last_pic_dir_name + '_DPI-list.xlsx'
    excel_filename = os.path.join(pic_dir, excel_filename)

    try:
        if os.path.exists(excel_filename):  # Delete Excel file if it exists
            os.remove(excel_filename)
    except PermissionError:
        messagebox.showerror("Error", "\n ******** Excel File Open! Please close it and run again.")
        sys.exit()

    with open(os.devnull, 'w') as devnull:
        subprocess.check_call(irfan_prog_run, stderr=devnull)

    # setup PROBLEM colors and bold for Excel
    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00',
                              end_color='FFFFFF00',
                              fill_type='solid')
    grey_fill = PatternFill(start_color='FF808080',
                            end_color='FF808080',
                            fill_type='solid')
    bold_font = Font(bold=True)
    italic_font = Font(italic=True)

    # ********* Extract data from TXT file

    # Setup Excel file
    excel_workbook = Workbook()
    excel_workbook.remove(excel_workbook['Sheet'])  # Remove default sheet
    zschriften_sheet = excel_workbook.create_sheet("DAI-Zeitschriften")
    reihen_sheet = excel_workbook.create_sheet("DAI-Reihen")
    interactive_sheet = excel_workbook.create_sheet("Max+Interactive")

    # Setup Headers
    header_to_col_Z = {
        'File name': 'A',
        'IMG Type & Compression': 'B',
        'Resolution': 'C',
        'Image Dim. (pixels)': 'D',
        'Image Orient.': 'E',
        'Print Size (CM)': 'F',
        'Print Size (IN)': 'G',
        '2 Sp. 4.03cm': 'H',
        '3 Sp. 6.28cm': 'I',
        '4 Sp. 8.52cm': 'J',
        '5 Sp. 10.76cm': 'K',
        '6 Sp. 13cm': 'L',
        '8 Sp. 17.5cm': 'M',
        'VolleS. 25.17cm': 'N',
    }

    header_to_col_R = {
        'File name': 'A',
        'IMG Type & Compression': 'B',
        'Resolution': 'C',
        'Image Dim. (pixels)': 'D',
        'Image Orient.': 'E',
        'Print Size (CM)': 'F',
        'Print Size (IN)': 'G',
        'A4 1 Sp. 7.75cm': 'H',
        'A4 2 Sp. 15.55cm': 'I',
        'A4 hoch 23.81cm': 'J',
        'Üformat 1 Sp. 8.775cm': 'K',
        'Üformat 2 Sp. 18.05cm': 'L',
        'Üformat hoch 26.9cm': 'M',
    }

    header_to_col_I = {
        'File name': 'A',
        'IMG Type & Compression': 'B',
        'Resolution': 'C',
        'Image Dim. (pixels)': 'D',
        'Image Orient.': 'E',
        'Print Size (CM)': 'F',
        'Print Size (IN)': 'G',
        'Max @ 300DPI': 'H',
        'Max @ 600DPI': 'I',
        '': 'J',
        'Goal cm': 'K',
        'DPI result': 'L',
    }

    # Write Headers for Excel DPI list sheet

    # Write Headers for Excel Zeitschriften sheet
    zschriften_sheet["A1"] = 'DPI List -- Data from IrfanView -- ' + datetime.datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S")
    excel_row = 4
    for header in header_to_col_Z:
        col = header_to_col_Z[header]
        zschriften_sheet[f'{col}{excel_row}'] = header
    for col in range(1, 15):  # 1-14 corresponds to columns A-N
        zschriften_sheet.cell(row=4, column=col).font = italic_font

    reihen_sheet["A1"] = 'DPI List -- Data from IrfanView -- ' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for header in header_to_col_R:
        col = header_to_col_R[header]
        reihen_sheet[f'{col}{excel_row}'] = header
    for col in range(1, 14):  # 1-13 corresponds to columns A-M
        reihen_sheet.cell(row=4, column=col).font = italic_font

    # Write Headers for Excel Interactive sheet
    interactive_sheet["A1"] = 'DPI List -- Data from IrfanView -- ' + datetime.datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S")
    for header in header_to_col_I:
        col = header_to_col_I[header]
        interactive_sheet[f'{col}{excel_row}'] = header
    for col in range(1, 13):  # 1-12 corresponds to columns A-L
        interactive_sheet.cell(row=4, column=col).font = italic_font

    excel_row = excel_row + 2

    # Parse TXT file and write to Excel
    numb_images = 0
    img_pix_x = img_pix_y = 0
    with open(irfan_info_txt, encoding='utf-16-le') as irfan_info_data:
        for line in irfan_info_data:

            if not line.strip():
                excel_row = excel_row + 1
                continue
            if not ' = ' in line:
                continue
            img_header, img_info = line.split(' = ')
            img_header = img_header.strip()
            img_info = img_info.strip()

            if img_header == 'File name':
                zschriften_sheet['A' + str(excel_row)] = img_info
                reihen_sheet['A' + str(excel_row)] = img_info
                interactive_sheet['A' + str(excel_row)] = img_info
                numb_images = numb_images + 1
                continue

            if img_header == 'Directory':
                zschriften_sheet['A2'] = img_info[:-1]
                reihen_sheet['A2'] = img_info[:-1]
                interactive_sheet['A2'] = img_info[:-1]
                continue

            if img_header == 'Compression':
                zschriften_sheet['B' + str(excel_row)] = img_info
                reihen_sheet['B' + str(excel_row)] = img_info
                interactive_sheet['B' + str(excel_row)] = img_info
                continue

            if img_header == 'Resolution':
                img_info = img_info.split(' DPI')[0]
                img_DPI_x, img_DPI_y = img_info.split(' x ')
                img_DPI_x = int(img_DPI_x)
                img_DPI_y = int(img_DPI_y)
                zschriften_sheet['C' + str(excel_row)] = str(img_DPI_x) + ' DPI'
                reihen_sheet['C' + str(excel_row)] = str(img_DPI_x) + ' DPI'
                interactive_sheet['C' + str(excel_row)] = str(img_DPI_x) + ' DPI'
                if not img_DPI_x == img_DPI_y:
                    reihen_sheet['C' + str(excel_row)].fill = red_fill
                if img_info in ['0 x 0', '96 x 96']:
                    reihen_sheet['A' + str(excel_row)].fill = grey_fill
                    interactive_sheet['A' + str(excel_row)].fill = grey_fill
                    zschriften_sheet['A' + str(excel_row)].fill = grey_fill
                    numb_images = numb_images - 1
                    img_info = img_header = img_pix = img_pix_x = img_pix_y = img_DPI_x = img_DPI_y = img_orient = img_landscape = img_coef = 0
                    continue
                continue

            if img_header == 'Image dimensions':
                img_pix = img_info.split('  Pixels')[0]
                img_pix_x, img_pix_y = map(int, img_pix.split(' x '))
                zschriften_sheet['D' + str(excel_row)] = img_pix.strip()
                reihen_sheet['D' + str(excel_row)] = img_pix.strip()
                interactive_sheet['D' + str(excel_row)] = img_pix.strip()
                continue

            if img_header == 'Print size':
                img_cm, img_in = img_info.split('; ')
                reihen_sheet['F' + str(excel_row)], zschriften_sheet['F' + str(excel_row)], interactive_sheet[
                    'F' + str(excel_row)] = img_cm, img_cm, img_cm
                reihen_sheet['G' + str(excel_row)], zschriften_sheet['G' + str(excel_row)], interactive_sheet[
                    'G' + str(excel_row)] = img_in, img_in, img_in
                img_in = img_in.split(' inches')[0]
                img_in_x, img_in_y = img_in.split(' x ')
                img_in_x = float(img_in_x)
                img_in_y = float(img_in_y)
                if img_in_x > img_in_y:
                    img_orient = 'Landscape'
                    img_landscape = True
                else:
                    img_orient = 'Portrait'
                    img_landscape = False
                zschriften_sheet['E' + str(excel_row)] = reihen_sheet['E' + str(excel_row)] = interactive_sheet[
                    'E' + str(excel_row)] = img_orient
                continue

            if img_header == 'Color depth':
                color_depth = float(img_info.split()[0].replace(',', '.'))
                if color_depth <= 2:
                    img_bitmap = True
                    reihen_sheet['B' + str(excel_row)] = zschriften_sheet['B' + str(excel_row)] = interactive_sheet[
                        'B' + str(excel_row)] = "BITMAP FILE"
                    reihen_sheet['B' + str(excel_row)].fill = zschriften_sheet['B' + str(excel_row)].fill = \
                    interactive_sheet['B' + str(excel_row)].fill = grey_fill
                else:
                    img_bitmap = False
                continue

            # ********* Calculate DPI for different print sizes. File date/time used bc last header
            if img_header == 'File date/time':

                # SETUP DPI TARGETS AND MATH

                ideal_targ_DPI = 600
                min_targ_DPI = 300
                ideal_targ_DPI_bit = 1600
                min_targ_DPI_bit = 1200

                def calculate_dpi_newx(img_pix_x, new_x_cm):
                    new_width_in = new_x_cm / 2.54  # convert cm to inches
                    result_DPI_fnct = round(img_pix_x / new_width_in)
                    return result_DPI_fnct

                def set_fill_color(result_DPI, excel_row, curr_column, curr_sheet):
                    if img_bitmap == False:
                        if result_DPI < min_targ_DPI:
                            curr_sheet[curr_column + str(excel_row)].fill = red_fill
                        elif min_targ_DPI <= result_DPI < ideal_targ_DPI:
                            curr_sheet[curr_column + str(excel_row)].fill = yellow_fill
                    elif img_bitmap == True:
                        if result_DPI < min_targ_DPI_bit:
                            curr_sheet[curr_column + str(excel_row)].fill = red_fill
                        elif min_targ_DPI_bit <= result_DPI < ideal_targ_DPI_bit:
                            curr_sheet[curr_column + str(excel_row)].fill = yellow_fill

                def calculate_max_widths(img_pix_x):
                    min_targ_DPI_width = round(((img_pix_x / min_targ_DPI) * 2.54), 1)
                    ideal_targ_DPI_width = round(((img_pix_x / ideal_targ_DPI) * 2.54), 1)
                    return min_targ_DPI_width, ideal_targ_DPI_width

                # ZEITSCHRIFTEN SHEET

                curr_sheet = zschriften_sheet

                # 2 Spalten 4.03cm
                new_x_cm = 4.03
                curr_column = 'H'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # 3 Spalten 6.28cm
                new_x_cm = 6.28
                curr_column = 'I'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # 4 Spalten 8.52cm
                new_x_cm = 8.52
                curr_column = 'J'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # 5 Spalten 10.76cm
                new_x_cm = 10.76
                curr_column = 'K'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # 6 Spalten 13cm
                new_x_cm = 13
                curr_column = 'L'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # 8 Spalten 17.5cm
                new_x_cm = 17.5
                curr_column = 'M'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # Volle Seite 25.17cm
                new_x_cm = 25.17
                curr_column = 'N'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # REIHEN SHEET

                curr_sheet = reihen_sheet

                # A4 1 Sp. 7.75cm
                new_x_cm = 7.75
                curr_column = 'H'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # A4 2 Sp. 15.55cm
                new_x_cm = 15.55
                curr_column = 'I'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # A4 hoch 23.81cm
                new_x_cm = 23.81
                curr_column = 'J'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # Üformat 1 Sp. 8.775cm
                new_x_cm = 8.775
                curr_column = 'K'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # Üformat 2 Sp. 18.05cm
                new_x_cm = 18.05
                curr_column = 'L'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # Üformat hoch 26.9cm
                new_x_cm = 26.9
                curr_column = 'M'
                result_DPI = calculate_dpi_newx(img_pix_x, new_x_cm)
                curr_sheet[curr_column + str(excel_row)] = result_DPI
                set_fill_color(result_DPI, excel_row, curr_column, curr_sheet)

                # INTERACTIVE SHEET

                curr_sheet = interactive_sheet

                # Max @ 400DPI & 800DPI
                min_targ_DPI_width, ideal_targ_DPI_width = calculate_max_widths(img_pix_x)
                curr_sheet['H' + str(excel_row)] = min_targ_DPI_width
                curr_sheet['I' + str(excel_row)] = ideal_targ_DPI_width

                # INPUT cm for DPI calculation
                curr_sheet['K' + str(excel_row)] = 25
                formula_cm_to_DPI = f"=ROUND({img_pix_x}/(K{excel_row}/2.54), 0)"
                curr_sheet[f'L{excel_row}'] = formula_cm_to_DPI

    # Calculate the number of images
    excel_row = excel_row + 1
    zschriften_sheet['A' + str(excel_row)] = reihen_sheet['A' + str(excel_row)] = interactive_sheet[
        'A' + str(excel_row)] = "Total Images: " + str(numb_images)
    zschriften_sheet['A' + str(excel_row)].font = reihen_sheet['A' + str(excel_row)].font = interactive_sheet[
        'A' + str(excel_row)].font = bold_font

    # ********* Format Excel Sheet

    # Set column widths and alignment
    def adjust_column_width(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    adjust_column_width(reihen_sheet)
    adjust_column_width(interactive_sheet)
    adjust_column_width(zschriften_sheet)

    reihen_sheet.column_dimensions['A'].width, reihen_sheet.column_dimensions['B'].width = 21, 21
    interactive_sheet.column_dimensions['A'].width, interactive_sheet.column_dimensions['B'].width = 21, 21
    zschriften_sheet.column_dimensions['A'].width, zschriften_sheet.column_dimensions['B'].width = 21, 21

# Column headers
    zschriften_sheet.merge_cells('H3:N3')
    zschriften_sheet[
        'H3'] = f"For fotos: under {min_targ_DPI} DPI red, {min_targ_DPI}-{ideal_targ_DPI} DPI yellow. Bitmap: under {min_targ_DPI_bit} DPI red, {min_targ_DPI_bit}-{ideal_targ_DPI_bit} DPI yellow."
    zschriften_sheet['H3'].font = bold_font
    zschriften_sheet['H3'].alignment = Alignment(horizontal='center')

    reihen_sheet.merge_cells('H3:M3')
    reihen_sheet[
        'H3'] = f"For fotos: under {min_targ_DPI} DPI red, {min_targ_DPI}-{ideal_targ_DPI} DPI yellow. Bitmap: under {min_targ_DPI_bit} DPI red, {min_targ_DPI_bit}-{ideal_targ_DPI_bit} DPI yellow."
    reihen_sheet['H3'].font = bold_font
    reihen_sheet['H3'].alignment = Alignment(horizontal='center')

    interactive_sheet.merge_cells('H3:I3')
    interactive_sheet['H3'] = "Max widths in cm"
    interactive_sheet['H3'].alignment = Alignment(horizontal='center')
    interactive_sheet.column_dimensions['F'].alignment = Alignment(horizontal='center')
    interactive_sheet.column_dimensions['L'].alignment = Alignment(horizontal='center')
    interactive_sheet['K3'] = "INPUT"
    interactive_sheet['L3'] = "OUTPUT"
    interactive_sheet['K3'].alignment = interactive_sheet['L3'].alignment = interactive_sheet['K4'].alignment = interactive_sheet['L4'].alignment = Alignment(horizontal='center')
    interactive_sheet['H3'].font = interactive_sheet['K3'].font = interactive_sheet['L3'].font = bold_font

# After creating all the sheets and before saving the workbook
    excel_workbook.active = zschriften_sheet

    excel_workbook.save(filename=excel_filename)
    excel_workbook.close()

# Remove TXT file

    if os.path.exists(irfan_info_txt):  # Delete TXT file if it already exists
        os.remove(irfan_info_txt)

def choose_directory():
    global pic_dir
    pic_dir = filedialog.askdirectory(title="Select Picture Folder")
    if not pic_dir:
        messagebox.showerror("Error", "No folder selected, please run again")
        return
    generate_excel()
    root.destroy()

def choose_fig_directories():
    global pic_dir
    parent_dir = filedialog.askdirectory(title="Select Parent Directory")
    if not parent_dir:
        messagebox.showerror("Error", "No folder selected, please run again")
        return
    subdirs = [os.path.normpath(os.path.join(parent_dir, d)) for d in os.listdir(parent_dir) if os.path.isdir(os.path.join(parent_dir, d))]
    for subdir in subdirs:
        if not os.listdir(subdir):
            continue
        pic_dir = subdir
        generate_excel()
    root.destroy()

# ********** GUI

try:
    check_version(__version__)
except Exception as e:
    messagebox.showerror("Error, could not check version", str(e))
    sys.exit(1)

try:
    get_irfanview_path()
except Exception as e:
    messagebox.showerror("Error, no IrfanView found", str(e))
    sys.exit(1)

root = tk.Tk()
root.title("Irfanview DPI Spreadsheet")
root.geometry("600x250")

# Icon
if getattr(sys, 'frozen', False):
    script_dir = sys._MEIPASS
else:
    script_dir = os.path.dirname(os.path.realpath(__file__))
icon_path = os.path.join(script_dir, 'icon', 'IrfanXcel.ico')
root.iconbitmap(icon_path)

# Styles
style = ttk.Style()
style.configure("BW.TLabel", foreground="#164194", font=("Arial", 25, "bold"))

style = ttk.Style()
style.configure("BW.TButton", font=("default", 14))

# Widgets
label = ttk.Label(root, text="Irfanview DPI Spreadsheet v0.9", style="BW.TLabel")
label.pack(pady=10)

button = ttk.Button(root, text="Choose directory with images: ", command=choose_directory, style="BW.TButton")
button.pack(pady=10)

button = ttk.Button(root, text="Or choose a directory to recursively process subdirectories: ", command=choose_fig_directories, style="BW.TButton")
button.pack(pady=10)

info_label2 = ttk.Label(root, text="This window will close when finished.\n", foreground="red", anchor='center')
info_label2.pack(side=tk.BOTTOM, fill='x')
info_label1 = ttk.Label(root, text="Please be patient, with many files or directories this program will take a while.", foreground="red", anchor='center')
info_label1.pack(side=tk.BOTTOM, fill='x')

root.mainloop()

# note for me on how to use pyinstaller. From terminal:
#       pyinstaller --clean --noconsole --icon=icon/IrfanXcel.ico --add-data "icon/IrfanXcel.ico;icon/" Irfanview_DPI_list.py
#  NO   pyinstaller --onefile --clean --noconsole --icon=icon/IrfanXcel.ico --add-data "icon/IrfanXcel.ico;icon/" Irfanview_DPI_list.py
