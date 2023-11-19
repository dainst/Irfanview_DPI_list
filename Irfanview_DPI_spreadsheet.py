#!/usr/bin/env python

__license__ = 'GPL'

# ********** Setup
import subprocess
import os
import sys
import datetime
import time
import requests
from distutils.spawn import find_executable
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
import tkinter as tk
from tkinter import filedialog

# ********** General Variables

# generate IDEAL and MIN Image Coefficient
# Formula: Image Quality Coefficient = ((Image width in inches * DPI) * (Image height in inches * DPI)) / 1000000
img_coef_page = 28.8  # this is the equivalent of a (8in x 10in @ 600DPI / 1000000) for a 1/1 page IDEAL
img_coef_page_min = 14.4  # this is the equivalent of a (8in x 10in @ 300DPI / 1000000) for a 1/1 page MIN

# Version number
script_ver_actual = '0.1'

# Terminal colors


def pr_blue(skk):
    print("\033[34m {}\033[00m" .format(skk))


def pr_red(skk):
    print("\033[91m {}\033[00m" .format(skk))


# Filenames
irfan_info_txt = 'DPI_list_irfanviewOUT.txt'
excel_filename = 'DPI_list.xlsx'

root = tk.Tk()
root.withdraw()
pic_dir = filedialog.askdirectory(title="Select Picture Folder")
pic_dir = pic_dir + '/'
if not pic_dir:
    sys.exit('No folder selected, please run again')

# ********** Intro
print()
pr_blue('                 ******************************************************')
pr_blue('                 *                    DAI ZWD-Redaktion               *')
pr_blue('                 *                Irfanview DPI Spreadsheet           *')
pr_blue('                 ******************************************************')
print()
pr_blue('All the info for this program comes from IrfanView, if that data is wrong then everything else will be wrong')
pr_blue('This is an executable, for the source code please look on GitHub for Irfanview_DPI_list')
pr_blue('The program will take a while, please be patient...')

# ********** Check for updates
# Get latest version from web
url = 'https://fabfab1.github.io/Irfanview_DPI_list/i_dpi_list_ver.html'
resp = requests.get(url)
script_ver_ideal = resp.text

# Compare versions
if script_ver_actual != script_ver_ideal:
    pr_red('A new version is available:'), print(script_ver_ideal)
    input("Press Enter to continue, or update the program")

# ********* Find and Run Irfanview
irfan_prog_name = 'i_view64.exe'
irfan_prog_cmd = find_executable(irfan_prog_name)
if not irfan_prog_cmd:
    irfan_prog_cmd = (os.path.expanduser('~')) + '\\PortableApps\\IrfanViewPortable\\App\\IrfanView64\\i_view64.exe'
if not irfan_prog_cmd:
    irfan_prog_name = 'i_view32.exe'
    irfan_prog_cmd = find_executable(irfan_prog_name)
if not irfan_prog_cmd:
    irfan_prog_cmd = (os.path.expanduser('~')) + '\\PortableApps\\IrfanViewPortable\\App\\IrfanView\\i_view32.exe'
if not irfan_prog_cmd:
    sys.exit('Irfanview not installed, please install and run again')

# This calls IrfanView and creates TXT file
irfan_info_txt = os.path.join(pic_dir, irfan_info_txt)
if os.path.exists(irfan_info_txt):  # Delete TXT file if it already exists
    os.remove(irfan_info_txt)
excel_filename = os.path.join(pic_dir, excel_filename)
if os.path.exists(excel_filename):  # Delete Excel file if it exists
    os.remove(excel_filename)
irfan_prog_cmd = irfan_prog_cmd + ' ' + '"' + pic_dir + '*.*' + '"' + ' /info=' + '"' + irfan_info_txt + '"'
# OLD REM subprocess.run(irfan_prog_cmd)
with open(os.devnull, 'w') as devnull:
    subprocess.check_call(irfan_prog_cmd, stderr=devnull)

# ********* Extract data from TXT file

# Setup Excel file
excel_workbook = Workbook()
excel_workbook.remove(excel_workbook['Sheet'])   # Remove default sheet
excel_sheet = excel_workbook.create_sheet("DPI list")
interactive_sheet = excel_workbook.create_sheet("Interactive")

# Setup Headers
header_to_col = {
    'File name': 'A',
    'IMG Type & Compression': 'B',
    'Resolution': 'C',
    'Image Dim. (pixels)': 'D',
    'Image Orient.': 'E',
    'Print Size (CM)': 'F',
    'Print Size (IN)': 'G',
    'Web Page': 'H',
    '1/4 Page': 'I',
    '1/2 Page': 'J',
    '1/1 Page': 'K',
    'Beilage 2.5x': 'L'
}

header_to_col2 = {
    'File name': 'A',
    'DPI': 'B',
    'Width CM': 'C',
    'Height CM': 'D',
    'Quality Coefficient': 'E',
    'Quality Enough?': 'F',
}

# Write Headers for Excel DPI list sheet
excel_sheet["A1"] = 'DPI List -- Data from IrfanView -- ' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
excel_row = 4
for header in header_to_col:
    col = header_to_col[header]
    excel_sheet[f'{col}{excel_row}'] = header

# Write Headers for Excel Interactive sheet
interactive_sheet["A1"] = 'DPI List -- Data from IrfanView -- ' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
for header in header_to_col2:
    col = header_to_col2[header]
    interactive_sheet[f'{col}{excel_row}'] = header
excel_row = excel_row + 2

# setup PROBLEM colors for Excel
red_fill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

# Parse TXT file and write to Excel
with open(irfan_info_txt, encoding='utf-16-le') as irfan_info_data:
    for line in irfan_info_data:

        if not line.strip():
            excel_row = excel_row + 1
            continue
        if not ' = ' in line:
            continue
        img_header, img_info = line.split(' = ')
        img_header = img_header.strip()
        img_info = img_info.strip()

        if img_header == 'File name':
            excel_sheet['A' + str(excel_row)] = img_info
            interactive_sheet['A' + str(excel_row)] = img_info

        if img_header == 'Directory':
            excel_sheet['A2'] = img_info

        if img_header == 'Compression':
            excel_sheet['B' + str(excel_row)] = img_info

        if img_header == 'Resolution':
            excel_sheet['C' + str(excel_row)] = img_info
            img_info = img_info.split(' DPI')[0]
            img_DPI_l, img_DPI_r = img_info.split(' x ')
            interactive_sheet['B' + str(excel_row)] = img_DPI_l
            if not img_DPI_l == img_DPI_r:
                excel_sheet['C' + str(excel_row)].fill = red_fill

        if img_header == 'Image dimensions':
            img_pix, img_orient = img_info.split(') (')
            img_pix = img_pix.split(' Pixels')[0]
            excel_sheet['D' + str(excel_row)] = img_pix.strip()
            img_orient = img_orient.split(')')[0]
            excel_sheet['E' + str(excel_row)] = img_orient

        if img_header == 'Print size':
            img_cm, img_in = img_info.split('; ')
            excel_sheet['F' + str(excel_row)] = img_cm
            excel_sheet['G' + str(excel_row)] = img_in
            img_in = img_in.split(' inches')[0]

            img_in_l, img_in_r = img_in.split(' x ')
            interactive_sheet['C' + str(excel_row)] = img_in_l
            interactive_sheet['D' + str(excel_row)] = img_in_r
            img_coef = (((float(img_in_l) * float(img_DPI_l)) * (float(img_in_r) * float(img_DPI_l)))/1000000)
            interactive_sheet['E' + str(excel_row)] = img_coef
            formula = '=IF(E{0}<$F$3,CONCATENATE("False"),CONCATENATE("True"))'.format(excel_row)
            interactive_sheet['F' + str(excel_row)].value = formula
            # Set conditional format rule to make "False" red/bold
            rule = FormulaRule(formula=['F1="False"'], fill=red_fill, font=Font(bold=True))
            interactive_sheet.conditional_formatting.add('F1:F5000', rule)

            # WEB Page
            excel_sheet['H3'] = (0.10 * img_coef_page)
            if img_coef >= (0.10 * img_coef_page):
                excel_sheet['H' + str(excel_row)] = True
            else:
                excel_sheet['H' + str(excel_row)] = False
                excel_sheet['H' + str(excel_row)].fill = red_fill

            # 1/4 Page
            excel_sheet['I3'] = (0.25 * img_coef_page)
            if img_coef >= (0.25 * img_coef_page):
                excel_sheet['I' + str(excel_row)] = True
            else:
                excel_sheet['I' + str(excel_row)] = False
                excel_sheet['I' + str(excel_row)].fill = red_fill

            # 1/2 Page
            excel_sheet['J3'] = (0.50 * img_coef_page)
            if img_coef >= (0.50 * img_coef_page):
                excel_sheet['J' + str(excel_row)] = True
            else:
                excel_sheet['J' + str(excel_row)] = False
                excel_sheet['J' + str(excel_row)].fill = red_fill

            # 1/1 Page
            excel_sheet['k3'] = (1.0 * img_coef_page)
            if img_coef >= (1.0 * img_coef_page):
                excel_sheet['K' + str(excel_row)] = True
            else:
                excel_sheet['K' + str(excel_row)] = False
                excel_sheet['K' + str(excel_row)].fill = red_fill

            # Beilage Page
            excel_sheet['L3'] = (2.50 * img_coef_page)
            if img_coef >= (2.5 * img_coef_page):
                excel_sheet['L' + str(excel_row)] = True
            else:
                excel_sheet['L' + str(excel_row)] = False
                excel_sheet['L' + str(excel_row)].fill = red_fill

# Now set up the interactive sheet
interactive_sheet['F3'] = 28.8
comment = """This is the minimum quality you are looking for. You can change this number 
and the info below changes. Generate this number using this formula: 
Image Quality Coefficient = ((Image width in inches * DPI) * (Image height in inches * DPI)) / 1000000"""
interactive_sheet['F3'].comment = Comment(comment, 'FAB')

# Set column widths and alignment
excel_sheet.column_dimensions['A'].width = 40
excel_sheet.column_dimensions['B'].width = 35
excel_sheet.column_dimensions['C'].width = 14
excel_sheet.column_dimensions['D'].width = 18
excel_sheet.column_dimensions['E'].width = 13
excel_sheet.column_dimensions['F'].width = 14
excel_sheet.column_dimensions['G'].width = 20
excel_sheet.column_dimensions['H'].width = 10
excel_sheet.column_dimensions['I'].width = 10
excel_sheet.column_dimensions['J'].width = 10
excel_sheet.column_dimensions['K'].width = 10
excel_sheet.column_dimensions['L'].width = 12
interactive_sheet.column_dimensions['A'].width = 40
interactive_sheet.column_dimensions['B'].width = 15
interactive_sheet.column_dimensions['C'].width = 14
interactive_sheet.column_dimensions['D'].width = 18
interactive_sheet.column_dimensions['E'].width = 18
interactive_sheet.column_dimensions['F'].width = 18
interactive_sheet.column_dimensions['F'].alignment = Alignment(horizontal='center')

excel_workbook.save(filename=excel_filename)
excel_workbook.close()

# ********* Outtro
if os.path.exists(irfan_info_txt):  # Delete TXT file if it already exists
    os.remove(irfan_info_txt)
print()
pr_blue('*****************************************')
print()
pr_blue('Done! Please check the Excel file.')
pr_blue('Remember that the image info is only as good as the info from IrfanView...')
pr_blue('so if authors "fudge" the image DPI then this program will be wrong!')
pr_blue('The Excel file has two sheets, the first is the DPI list and the second is interactive.')
pr_blue('hope this was helpful.')
time.sleep(10)
